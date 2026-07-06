"""
<<<<<<< HEAD
prepare.py — normalizes a spreadsheet for clean PDF printing.
=======
prepare.py — normalizes a spreadsheet so it prints as a clean table.
>>>>>>> 3d4cafc43ce1987aa97b4da38fbbe6f7845c4cab

Usage:
    python3 prepare.py csv   <input.csv>   <output.xlsx>
    python3 prepare.py style <input.xlsx>  <output.xlsx>

<<<<<<< HEAD
"csv" mode reads a plain CSV (which has no formatting at all) and builds
a fully styled table: borders, bold header row, landscape + fit-to-width.

"style" mode opens an EXISTING workbook — one that may already have its
own design (colors, merged cells, custom borders, an invoice layout,
etc.) — and deliberately leaves that design alone. It only:
  - trims empty trailing/leading rows and columns (fixes stray blank
    pages caused by formatting applied far beyond the real data)
  - sets fit-to-width so columns don't get split across separate
    "page groups"
It does NOT add borders, fill colors, or force an orientation, because
doing so would overwrite formatting the file's author already chose.
=======
"csv" mode reads a CSV and writes a styled workbook.
"style" mode opens an existing workbook and applies the same table
styling + print setup, without touching the cell values.
>>>>>>> 3d4cafc43ce1987aa97b4da38fbbe6f7845c4cab
"""

import sys
import os
import csv
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


<<<<<<< HEAD
def find_used_bounds(ws):
    """Return (first_row, last_row, first_col, last_col) that actually
    contain non-empty values, or (None, None, None, None) if the sheet
    has no data at all."""
    first_row = last_row = first_col = last_col = None
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is not None and str(cell.value).strip() != '':
                r, c = cell.row, cell.column
                if first_row is None or r < first_row:
                    first_row = r
                if last_row is None or r > last_row:
                    last_row = r
                if first_col is None or c < first_col:
                    first_col = c
                if last_col is None or c > last_col:
                    last_col = c
    return first_row, last_row, first_col, last_col


def trim_to_used_range(ws):
    """Delete empty trailing AND leading rows/columns so the sheet's
    used range matches the real data exactly. Returns (max_row, max_col)
    of the trimmed sheet, or (0, 0) if the sheet was entirely empty."""
    first_row, last_row, first_col, last_col = find_used_bounds(ws)

    if first_row is None:
        return 0, 0  # sheet is completely empty — leave it alone

    if ws.max_row > last_row:
        ws.delete_rows(last_row + 1, ws.max_row - last_row)
    if ws.max_column > last_col:
        ws.delete_cols(last_col + 1, ws.max_column - last_col)

    if first_row > 1:
        ws.delete_rows(1, first_row - 1)
        last_row -= (first_row - 1)
    if first_col > 1:
        ws.delete_cols(1, first_col - 1)
        last_col -= (first_col - 1)

    return last_row, last_col


def apply_print_setup_only(ws, max_row, max_col):
    """Fix pagination without touching visual formatting."""
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.print_title_rows = '1:1'
    last_col_letter = get_column_letter(max_col)
    ws.print_area = f"A1:{last_col_letter}{max_row}"

    # Explicitly turn OFF printed gridlines. Some exporters print
    # gridlines across the whole used range by default unless this is
    # set — which looks like a grid of empty boxes around content that
    # was never meant to have any lines at all.
    ws.print_options.gridLines = False
    ws.print_options.gridLinesSet = True
    ws.sheet_view.showGridLines = False


def style_and_setup_fully(ws, max_row, max_col):
    """Full generated-table look: borders, header row, landscape, fit-to-width.
    Used only for plain CSV imports, which have no design to preserve."""
=======
def style_workbook(wb):
>>>>>>> 3d4cafc43ce1987aa97b4da38fbbe6f7845c4cab
    thin = Side(style='thin', color='999999')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill = PatternFill(start_color='E8EFEA', end_color='E8EFEA', fill_type='solid')
    header_font = Font(bold=True)

<<<<<<< HEAD
    for row in ws.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col):
        for cell in row:
            cell.border = border

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font

    for col_idx in range(1, max_col + 1):
        col_letter = get_column_letter(col_idx)
        max_len = 0
        scan_rows = min(max_row, 300)
        for row_idx in range(1, scan_rows + 1):
            val = ws.cell(row=row_idx, column=col_idx).value
            if val is not None:
                max_len = max(max_len, len(str(val)))
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 8), 40)

    ws.page_setup.orientation = 'landscape'
    apply_print_setup_only(ws, max_row, max_col)
=======
    for ws in wb.worksheets:
        max_row = ws.max_row
        max_col = ws.max_column
        if max_row == 0 or max_col == 0:
            continue

        # Real borders on every used cell, so the PDF always shows gridlines
        # (relying on the "print gridlines" checkbox is inconsistent across apps).
        for row in ws.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col):
            for cell in row:
                cell.border = border

        # Bold, shaded header row.
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font

        # Rough auto-width so columns aren't cramped or absurdly wide.
        for col_idx in range(1, max_col + 1):
            col_letter = get_column_letter(col_idx)
            max_len = 0
            scan_rows = min(max_row, 300)
            for row_idx in range(1, scan_rows + 1):
                val = ws.cell(row=row_idx, column=col_idx).value
                if val is not None:
                    max_len = max(max_len, len(str(val)))
            ws.column_dimensions[col_letter].width = min(max(max_len + 2, 8), 40)

        # Page setup: force every column onto one page width, let rows
        # flow across as many pages tall as needed, and repeat the header.
        ws.page_setup.orientation = 'landscape'
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0
        ws.sheet_properties.pageSetUpPr.fitToPage = True
        ws.print_title_rows = '1:1'

        last_col_letter = get_column_letter(max_col)
        ws.print_area = f"A1:{last_col_letter}{max_row}"
>>>>>>> 3d4cafc43ce1987aa97b4da38fbbe6f7845c4cab


def convert_csv_to_xlsx(csv_path, xlsx_path):
    wb = Workbook()
    ws = wb.active
    with open(csv_path, newline='', encoding='utf-8-sig', errors='replace') as f:
        reader = csv.reader(f)
        for row in reader:
            ws.append(row)
<<<<<<< HEAD

    for ws_ in wb.worksheets:
        max_row, max_col = trim_to_used_range(ws_)
        if max_row == 0 or max_col == 0:
            continue
        style_and_setup_fully(ws_, max_row, max_col)

=======
    style_workbook(wb)
>>>>>>> 3d4cafc43ce1987aa97b4da38fbbe6f7845c4cab
    wb.save(xlsx_path)


def restyle_xlsx(input_path, output_path):
<<<<<<< HEAD
    """Preserve the file's existing design; only fix pagination."""
    wb = load_workbook(input_path)

    for ws in wb.worksheets:
        max_row, max_col = trim_to_used_range(ws)
        if max_row == 0 or max_col == 0:
            continue
        apply_print_setup_only(ws, max_row, max_col)

=======
    wb = load_workbook(input_path)
    style_workbook(wb)
>>>>>>> 3d4cafc43ce1987aa97b4da38fbbe6f7845c4cab
    wb.save(output_path)


def main():
    if len(sys.argv) != 4:
        print('Usage: prepare.py <csv|style> <input> <output>', file=sys.stderr)
        sys.exit(1)

    mode, input_path, output_path = sys.argv[1], sys.argv[2], sys.argv[3]

    if mode == 'csv':
        convert_csv_to_xlsx(input_path, output_path)
    elif mode == 'style':
        restyle_xlsx(input_path, output_path)
    else:
        print(f'Unknown mode: {mode}', file=sys.stderr)
        sys.exit(1)


if __name__ == '__main__':
    main()
